import phonenumbers
import openpyxl
import os
import pandas as pd

# GET FILES
os.chdir('C:\\Users\\CSR001\\Documents\\CleanLeads')
# FILE TO CLEAN
Originwb = openpyxl.load_workbook('Lead.xlsx')
Originsheet = Originwb['Sheet2']

# FILES FOR CLEAN LEADS
Daywb = openpyxl.load_workbook('CleanLeads.xlsx')
Daysheet = Daywb['Sheet1']
# DAY
# NIGHT
Nightwb = openpyxl.load_workbook('CleanLeads.xlsx')
Nightsheet = Nightwb['Sheet1']
# AFRICA
Africawb = openpyxl.load_workbook('CleanLeads.xlsx')
Africasheet = Africawb['Sheet1']
# NA
NAwb = openpyxl.load_workbook('CleanLeads.xlsx')
NAsheet = NAwb['Sheet1']
# SPAIN
Spainwb = openpyxl.load_workbook('CleanLeads.xlsx')
Spainsheet = Spainwb['Sheet1']
# ITALY
Italywb = openpyxl.load_workbook('CleanLeads.xlsx')
Italysheet = Italywb['Sheet1']
# BRAZIL
Brazilwb = openpyxl.load_workbook('CleanLeads.xlsx')
Brazilsheet = Brazilwb['Sheet1']

# VARIABLES
NotCleanRows = Originsheet.max_row  # MAX ROWS OF LEADS TO BE CLEANED

NotCleanEmail = []  # ARRAYS FOR UNCLEAN LEADS
NotCleanNumbers = []
NotCleanFname = []
NotCleanLname = []
NotCleanArea = []

CleanEmail = []  # ARRAYS FOR CLEAN LEADS
CleanNumbers = []
CleanFname = []
CleanLname = []
CleanArea = []

# FIRST CLEAN
for i in range(2, NotCleanRows + 1):
    NCRemail = Originsheet.cell(i, 1)
    NCRfname = Originsheet.cell(i, 2)
    NCRlname = Originsheet.cell(i, 3)
    NCRnum = Originsheet.cell(i, 4)
    NCRarea = Originsheet.cell(i, 5)
    if NCRarea is None:
        try:
            NumParsed = phonenumbers.parse("+" + str(NCRnum.value), None)
            NotCleanNumbers.append(NumParsed)
            NotCleanEmail.append(NCRemail.value)
            NotCleanFname.append(NCRfname.value)
            NotCleanLname.append(NCRlname.value)
            NotCleanArea.append(NCRarea.value)
        except:
            pass
    else:
        try:
            x = phonenumbers.parse("+" + str(NCRnum.value), str(NCRarea.value))
            NotCleanNumbers.append(x)
            NotCleanEmail.append(NCRemail.value)
            NotCleanFname.append(NCRfname.value)
            NotCleanLname.append(NCRlname.value)
            NotCleanArea.append(NCRarea.value)
        except:
            pass
# SECOND CLEAN
RowsClean = 0
for j in NotCleanNumbers:
    if phonenumbers.is_possible_number(j):
        if phonenumbers.is_valid_number(j):
            CleanNumbers.append(j)
            CleanEmail.append(NotCleanEmail[RowsClean])
            CleanFname.append(NotCleanFname[RowsClean])
            CleanLname.append(NotCleanLname[RowsClean])
            CleanArea.append(NotCleanArea[RowsClean])
    RowsClean += 1
# SEPARATE COUNTRIES
# ROWS TO INSERT IN EXCEL
RowNA = 1
RowBR = 1
RowAF = 1
RowIT = 1
RowSP = 1

TotalAS = 0
TotalGCC = 0
# ARRAYS FOR CLEANED DATA FOR DIFFERENT COUNTRIES
Array = 0
# ASIA CLEAN LEADS
for CCode in CleanNumbers:
    CountryInitial = phonenumbers.region_code_for_number(CCode)
    if CountryInitial in ['HK', 'SG', 'ID', 'JP', 'MO', 'MY', 'KR', 'TW']:
        TotalAS += 1
    # GCC CLEAN LEADS
    if CountryInitial in ['BH', 'KW', 'QA', 'SA', 'AE', 'LB']:
        TotalGCC += 1
    # NORTH AMERICA CLEAN LEADS
    if CountryInitial in ['CA', 'TT', 'BS']:
        RowNA += 1
        EmailCel = NAsheet.cell(RowNA, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = NAsheet.cell(RowNA, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = NAsheet.cell(RowNA, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = NAsheet.cell(RowNA, 4)
        LnameCel.value = str(CleanLname[Array])
    # ITALY CLEAN LEADS
    if CountryInitial in ['IT']:
        RowIT += 1
        EmailCel = Italysheet.cell(RowIT, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = Italysheet.cell(RowIT, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = Italysheet.cell(RowIT, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = Italysheet.cell(RowIT, 4)
        LnameCel.value = str(CleanLname[Array])
    # SPAIN CLEAN LEADS
    if CountryInitial in ['ES']:
        RowSP += 1
        EmailCel = Spainsheet.cell(RowSP, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = Spainsheet.cell(RowSP, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = Spainsheet.cell(RowSP, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = Spainsheet.cell(RowSP, 4)
        LnameCel.value = str(CleanLname[Array])
    # AFRICA CLEAN LEADS
    if CountryInitial in ['MG', 'ZA']:
        RowAF += 1
        EmailCel = Africasheet.cell(RowAF, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = Africasheet.cell(RowAF, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = Africasheet.cell(RowAF, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = Africasheet.cell(RowAF, 4)
        LnameCel.value = str(CleanLname[Array])
    # BRAZIL CLEAN LEADS
    if CountryInitial in ['BR']:
        RowBR += 1
        EmailCel = Brazilsheet.cell(RowBR, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = Brazilsheet.cell(RowBR, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = Brazilsheet.cell(RowBR, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = Brazilsheet.cell(RowBR, 4)
        LnameCel.value = str(CleanLname[Array])
    Array += 1
# DAY/NIGHT SHIFT LEADS
RowDay = 1
RowNight = 1
AsiaRows = 1
GCCRows = 1
Array = 0
DayAsia = int(TotalAS * 0.45)
DayGCC = int(TotalGCC * 0.35)
print(DayAsia)
print(DayGCC)
for CCode in CleanNumbers:
    CountryInitial = phonenumbers.region_code_for_number(CCode)
    if CountryInitial in ['HK', 'SG', 'ID', 'JP', 'MO', 'MY', 'KR', 'TW']:
        if AsiaRows <= DayAsia:
            print('ASIA DAY')
            EmailCel = Daysheet.cell(RowDay, 1)
            EmailCel.value = str(CleanEmail[Array])
            NumCel = Daysheet.cell(RowDay, 2)
            NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
            NumCel.number_format = '0'
            FnameCel = Daysheet.cell(RowDay, 3)
            FnameCel.value = str(CleanFname[Array])
            LnameCel = Daysheet.cell(RowDay, 4)
            LnameCel.value = str(CleanLname[Array])
            ACel = Daysheet.cell(RowDay, 5)
            ACel.value = '=RAND()'
            RowDay += 1
            AsiaRows += 1
        else:
            print('ASIA NIGHT')
            EmailCel = Nightsheet.cell(RowNight, 1)
            EmailCel.value = str(CleanEmail[Array])
            NumCel = Nightsheet.cell(RowNight, 2)
            NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
            NumCel.number_format = '0'
            FnameCel = Nightsheet.cell(RowNight, 3)
            FnameCel.value = str(CleanFname[Array])
            LnameCel = Nightsheet.cell(RowNight, 4)
            LnameCel.value = str(CleanLname[Array])
            ACel = Nightsheet.cell(RowNight, 5)
            ACel.value = '=RAND()'
            RowNight += 1
            AsiaRows += 1
    # GCC CLEAN LEADS
    if CountryInitial in ['BH', 'KW', 'QA', 'SA', 'AE', 'LB']:
        if GCCRows <= DayGCC:
            print('GCCDAY')
            EmailCel = Daysheet.cell(RowDay, 1)
            EmailCel.value = str(CleanEmail[Array])
            NumCel = Daysheet.cell(RowDay, 2)
            NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
            NumCel.number_format = '0'
            FnameCel = Daysheet.cell(RowDay, 3)
            FnameCel.value = str(CleanFname[Array])
            LnameCel = Daysheet.cell(RowDay, 4)
            LnameCel.value = str(CleanLname[Array])
            ACel = Daysheet.cell(RowDay, 5)
            ACel.value = '=RAND()'
            RowDay += 1
            GCCRows += 1
        else:
            print('GCC NIGHT')
            EmailCel = Nightsheet.cell(RowNight, 1)
            EmailCel.value = str(CleanEmail[Array])
            NumCel = Nightsheet.cell(RowNight, 2)
            NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
            NumCel.number_format = '0'
            FnameCel = Nightsheet.cell(RowNight, 3)
            FnameCel.value = str(CleanFname[Array])
            LnameCel = Nightsheet.cell(RowNight, 4)
            LnameCel.value = str(CleanLname[Array])
            ACel = Nightsheet.cell(RowNight, 5)
            ACel.value = '=RAND()'
            RowNight += 1
            GCCRows += 1
    # OCEANIA CLEAN LEADS
    if CountryInitial in ['NZ', 'AU']:
        EmailCel = Daysheet.cell(RowDay, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = Daysheet.cell(RowDay, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = Daysheet.cell(RowDay, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = Daysheet.cell(RowDay, 4)
        LnameCel.value = str(CleanLname[Array])
        ACel = Daysheet.cell(RowDay, 5)
        ACel.value = '=RAND()'
        RowDay += 1
    # EUROPE CLEAN LEADS
    if CountryInitial in ['LU', 'LI', 'IE', 'IS', 'NL', 'NO', 'PL', 'SE', 'CH', 'GB', 'DK', 'FI', 'DE',
                          'BG', 'HR', 'CY', 'EE', 'GR', 'HU', 'IM', 'LT', 'MT', 'MC', 'RO', 'CZ', 'PT']:
        EmailCel = Nightsheet.cell(RowNight, 1)
        EmailCel.value = str(CleanEmail[Array])
        NumCel = Nightsheet.cell(RowNight, 2)
        NumCel.value = int(str(CCode.country_code) + str(CCode.national_number))
        NumCel.number_format = '0'
        FnameCel = Nightsheet.cell(RowNight, 3)
        FnameCel.value = str(CleanFname[Array])
        LnameCel = Nightsheet.cell(RowNight, 4)
        LnameCel.value = str(CleanLname[Array])
        ACel = Nightsheet.cell(RowNight, 5)
        ACel.value = '=RAND()'
        RowNight += 1
    Array += 1

Brazilwb.save('Leads\\TestCleanerBrazil.xlsx')
Africawb.save('Leads\\TestCleanerAfrica.xlsx')
Spainwb.save('Leads\\TestCleanerSpain.xlsx')
Italywb.save('Leads\\TestCleanerItaly.xlsx')
NAwb.save('Leads\\TestCleanerNorthA.xlsx')
Daywb.save('Leads\\TestCleanerDay.xlsx')
Nightwb.save('Leads\\TestCleanerNight.xlsx')